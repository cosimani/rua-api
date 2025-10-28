from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session, joinedload, aliased
from database.config import get_db  # Importá get_db desde config.py

from models.users import User, UserGroup, Group
from models.proyecto import Proyecto
from models.carpeta import Carpeta, DetalleNNAEnCarpeta, DetalleProyectosEnCarpeta
from models.nna import Nna
from models.convocatorias import DetalleNNAEnConvocatoria


from helpers.utils import parse_date
from models.eventos_y_configs import RuaEvento
from security.security import get_current_user, verify_api_key, require_roles

from sqlalchemy.sql import func, or_

from fastapi.responses import FileResponse
from helpers.utils import EstadisticasPDF, calcular_estadisticas_generales

from tempfile import NamedTemporaryFile
from datetime import date, datetime
from typing import List, Dict, Any, Optional

from models.nna import Nna
from models.convocatorias import Convocatoria

from sqlalchemy import text, func, distinct, or_, select, and_

import os
from fastapi import BackgroundTasks
from database.config import SessionLocal  # para obtener engine/bind
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from starlette.concurrency import run_in_threadpool

from helpers.utils import (
    JOBSTORE_EXPORT_DIR,
    jobstore_create_job,
    jobstore_update_job,
    jobstore_read_job,
)


estadisticas_router = APIRouter()


def g(stats: dict, path: str, default=0):
    """
    Lee claves anidadas con path tipo 'proyectos.resumen.proyectos_viables'.
    Si no existe, devuelve default.
    """
    cur = stats
    try:
        for k in path.split("."):
            cur = cur[k]
        return cur
    except Exception:
        return default
    


@estadisticas_router.get("/generales", response_model=dict, 
                         dependencies=[Depends( verify_api_key ), 
                                       Depends(require_roles(["administrador", "supervision", "supervisora", "profesional", "coordinadora"]))])
def get_estadisticas(db: Session = Depends(get_db)):
    return calcular_estadisticas_generales(db)


@estadisticas_router.get("/estadisticas-portada", response_model=dict,
    dependencies=[ Depends(verify_api_key),
        Depends(require_roles(["administrador","supervision","supervisora","profesional","coordinadora"]))])
def get_estadisticas_portada(db: Session = Depends(get_db)):
    try:
        # 1) Proyectos viables (solo RUA)
        proyectos_viables = (
            db.query(Proyecto)
              .filter(Proyecto.ingreso_por == 'rua',
                      Proyecto.estado_general == 'viable')
              .count()
        )

        # 2) Proyectos en entrevistas (RUA + convocatoria)
        proyectos_en_entrevistas = (
            db.query(Proyecto)
              .filter(
                  Proyecto.estado_general == 'entrevistando',
                  Proyecto.ingreso_por.in_(('rua', 'convocatoria'))
              )
              .count()
        )

        # 3) NNAs en guarda (provisoria o confirmada) – contar DISTINCT nna_id
        nna_en_guarda = (
            db.query(func.count(distinct(DetalleNNAEnCarpeta.nna_id)))
              .join(DetalleProyectosEnCarpeta,
                    DetalleProyectosEnCarpeta.carpeta_id == DetalleNNAEnCarpeta.carpeta_id)
              .join(Proyecto, Proyecto.proyecto_id == DetalleProyectosEnCarpeta.proyecto_id)
              .filter(Proyecto.estado_general.in_(('guarda_provisoria','guarda_confirmada')))
              .scalar()
        ) or 0

        # 4) NNAs en adopción definitiva – DISTINCT nna_id
        nna_en_adopcion_definitiva = (
            db.query(func.count(distinct(DetalleNNAEnCarpeta.nna_id)))
              .join(DetalleProyectosEnCarpeta,
                    DetalleProyectosEnCarpeta.carpeta_id == DetalleNNAEnCarpeta.carpeta_id)
              .join(Proyecto, Proyecto.proyecto_id == DetalleProyectosEnCarpeta.proyecto_id)
              .filter(Proyecto.estado_general == 'adopcion_definitiva')
              .scalar()
        ) or 0

        return {
            "proyectos_viables": proyectos_viables,
            "proyectos_en_entrevistas": proyectos_en_entrevistas,
            "nna_en_guarda": nna_en_guarda,
            "nna_en_adopcion_definitiva": nna_en_adopcion_definitiva,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



@estadisticas_router.get("/informe_general", dependencies=[ Depends(verify_api_key),
        Depends(require_roles(["administrador", "supervision", "supervisora", "coordinadora"]))],)
def generar_pdf_estadisticas(db: Session = Depends(get_db)):
    stats = calcular_estadisticas_generales(db)

    pdf = EstadisticasPDF()
    pdf.add_page()

    # =========================================================
    # RESUMEN GENERAL
    # =========================================================
    pdf.section_title("Resumen General de Indicadores Clave")

    resumen_general = [
        ["Indicador", "Cantidad"],
        ["Proyectos viables disponibles", g(stats, "proyectos.resumen.proyectos_viables")],
        ["Proyectos en entrevistas (ingresado por RUA)", g(stats, "proyectos.resumen.proyectos_en_entrevistas_rua")],
        ["Proyectos en entrevistas (por convocatoria)", g(stats, "proyectos.resumen.proyectos_en_entrevistas_convocatoria")],
        ["NNAs en Guarda Provisoria", g(stats, "nna.nna_en_guarda_provisoria")],
        ["NNAs en Guarda Confirmada", g(stats, "nna.nna_en_guarda_confirmada")],
        ["NNAs con Adopción Definitiva", g(stats, "nna.nna_en_adopcion_definitiva")],
        ["Postulaciones totales (todas las convocatorias)", g(stats, "usuarios.postulaciones_totales")],  # ← NUEVO
    ]
    pdf.add_table(resumen_general)

    # =========================================================
    # 1) PROYECTOS
    # =========================================================
    pdf.section_title("1) Estadísticas de Proyectos ingresados por RUA")

    tabla_1 = [
        ["Tipo", "Presentando doc.", "En revisión", "Calendarizables", "Entrevistando"],
        [
            "Pareja",
            g(stats, "proyectos.tipos.en_pareja_subiendo_documentacion"),
            g(stats, "proyectos.tipos.en_pareja_en_revision"),
            g(stats, "proyectos.tipos.en_pareja_aprobados_para_calendarizar"),
            g(stats, "proyectos.tipos.entrevistando_en_pareja"),
        ],
        [
            "Monoparental",
            g(stats, "proyectos.tipos.monoparentales_subiendo_documentacion"),
            g(stats, "proyectos.tipos.monoparentales_en_revision"),
            g(stats, "proyectos.tipos.monoparentales_aprobados_para_calendarizar"),
            g(stats, "proyectos.tipos.entrevistando_monoparental"),
        ],
    ]
    pdf.add_table(tabla_1)

    tabla_2 = [
        ["Tipo", "En suspenso", "No viables", "Baja definitiva"],
        [
            "Pareja",
            g(stats, "proyectos.tipos.en_pareja_en_suspenso"),  # total
            g(stats, "proyectos.tipos.en_pareja_no_viables"),   # total
            g(stats, "proyectos.tipos.en_pareja_baja_definitiva"),  # ⚠️ ver nota abajo
        ],
        [
            "Monoparental",
            g(stats, "proyectos.tipos.monoparentales_en_suspenso"),  # total
            g(stats, "proyectos.tipos.monoparentales_no_viables"),   # total
            g(stats, "proyectos.tipos.monoparentales_baja_definitiva"),  # ⚠️ ver nota abajo
        ],
    ]
    pdf.add_table(tabla_2)

    tabla_3 = [
        ["Tipo", "Viables disponibles", "Adopción definitiva"],
        [
            "Pareja",
            g(stats, "proyectos.tipos.proyectos_en_pareja_viable"),  # ← corregido
            g(stats, "proyectos.tipos.proyectos_adopcion_definitiva_pareja"),
        ],
        [
            "Monoparental",
            g(stats, "proyectos.tipos.proyectos_monoparental_viable"),
            g(stats, "proyectos.tipos.proyectos_adopcion_definitiva_monoparental"),
        ],
    ]
    pdf.add_table(tabla_3)

    tabla_4 = [
        ["Ingreso por", "Cantidad"],
        ["RUA", g(stats, "proyectos.por_ingreso.rua")],
        ["Oficio", g(stats, "proyectos.por_ingreso.oficio")],
        ["Convocatoria", g(stats, "proyectos.por_ingreso.convocatoria")],
    ]
    pdf.add_table(tabla_4)

    # (Opcional) Proyectos por estado (muestra principales)
    por_estado = g(stats, "proyectos.por_estado", {})
    if isinstance(por_estado, dict) and por_estado:
        pdf.section_title("Proyectos por estado (principales)")
        # Ordeno por cantidad desc y muestro top 10
        top = sorted(por_estado.items(), key=lambda x: x[1], reverse=True)[:10]
        tabla_estados = [["Estado", "Cantidad"]] + [[k, v] for k, v in top]
        pdf.add_table(tabla_estados)

    pdf.add_page()

    # =========================================================
    # 2) NNA
    # =========================================================

    pdf.section_title("2) NNA")
    tabla_nna_clip = [
        ["Indicador", "Cantidad"],
        ["NNAs en Proyecto con Guarda Provisoria", g(stats, "nna.nna_en_guarda_provisoria")],
        ["NNAs en Proyecto con Guarda Confirmada", g(stats, "nna.nna_en_guarda_confirmada")],
        ["NNAs en Proyecto con Adopción Definitiva", g(stats, "nna.nna_en_adopcion_definitiva")],
    ]
    pdf.add_table(tabla_nna_clip)
    
    # NNA: distribución por estado
    nna_por_estado = g(stats, "nna.por_estado", {})
    if isinstance(nna_por_estado, dict) and nna_por_estado:
        pdf.section_title("NNA por estado")
        tabla_nna_estado = [["Estado", "Cantidad"]] + [[k, v] for k, v in nna_por_estado.items()]
        pdf.add_table(tabla_nna_estado)

    # NNA: distribución por edades
    edades = g(stats, "nna.edades", {})
    if isinstance(edades, dict) and edades:
        pdf.section_title("Distribución de NNA por edades (todos los estados)")
        tabla_edades = [
            ["Rango", "Cantidad"],
            ["0-6", edades.get("0_6", 0)],
            ["7-11", edades.get("7_11", 0)],
            ["12-17", edades.get("12_17", 0)],
            ["18+", edades.get("18_mas", 0)],
        ]
        pdf.add_table(tabla_edades)

    pdf.add_page()

    # =========================================================
    # 3) PRETENSOS
    # =========================================================
    pdf.section_title("3) Estadísticas de Pretensos")

    pretensos_data = [
        ["Indicador", "Cantidad"],
        ["Activos (Pretensos con clave y activación con mail)", g(stats, "usuarios.usuarios_totales")],  # usuarios_totales redefinido
        ["Usuarios inactivos (aún no activaron con mail)", g(stats, "usuarios.sin_activar")],
        ["Con Curso y DDJJ firmada", g(stats, "usuarios.con_curso_con_ddjj")],
        
        ["Con documentación aprobada pero sin proyecto", g(stats, "usuarios.usuarios_sin_proyecto")],
        ["Postulados y CON usuario activo en RUA", g(stats, "usuarios.usuarios_postulados_y_rua")],            # ← NUEVO
        ["Postulados y SIN usuario en RUA (sólo se postularon)", g(stats, "usuarios.usuarios_postulados_y_no_rua")],        # ← NUEVO
        ["Postulaciones (incluyendo pretensos en varias convocatorias)", g(stats, "usuarios.postulaciones_totales")],                   # ← NUEVO
    ]
    pdf.add_table(pretensos_data)

    # DDJJ
    pdf.section_title("DDJJ (flexibilidades y condiciones)")
    ddjj = g(stats, "ddjj", {})
    tabla_ddjj = [
        ["Indicador", "Cantidad"],
        ["DDJJ firmadas", ddjj.get("firmadas", 0)],
        ["DDJJ no firmadas", ddjj.get("no_firmadas", 0)],
        ["Flexibilidad de edad (alguna)", ddjj.get("flex_edad_alguna", 0)],
        ["Aceptan discapacidad", ddjj.get("acepta_discapacidad", 0)],
        ["Aceptan enfermedades", ddjj.get("acepta_enfermedad", 0)],
        ["Aceptan grupo de hermanos", ddjj.get("acepta_grupo_hermanos", 0)],
    ]
    pdf.add_table(tabla_ddjj)

    pdf.add_page()

    # =========================================================
    # 4) TIEMPOS
    # =========================================================
    pdf.section_title("4) Tiempos (Proyectos y Pretensos)")

    # Proyectos: promedio de días por estado
    tiempos_proy = g(stats, "tiempos.proyectos", {})
    prom_por_estado = tiempos_proy.get("promedio_dias_por_estado", {})
    if isinstance(prom_por_estado, dict) and prom_por_estado:
        pdf.section_title("Tiempos de proyectos por estado (promedio en días)")
        tabla_tiempos_estado = [["Estado", "Promedio días"]] + [
            [k, round(v, 2)] for k, v in sorted(prom_por_estado.items(), key=lambda x: x[0])
        ]
        pdf.add_table(tabla_tiempos_estado)

    # Proyectos: tiempo total promedio por proyecto
    pdf.add_table([
        ["Métrica", "Valor"],
        ["Promedio de días por proyecto (ciclo completo)", round(g(stats, "tiempos.proyectos.promedio_dias_total_por_proyecto", 0), 2)]
    ])

    # Pretensos: pipeline de tiempos
    tiempos_pret = g(stats, "tiempos.pretensos", {})
    pdf.section_title("Pipeline de pretensos (promedios en días)")
    tabla_pret = [
        ["Tramo", "Promedio días"],
        ["Curso aprobado -> DDJJ firmada", round(tiempos_pret.get("promedio_dias_curso_a_ddjj", 0), 2)],
        ["DDJJ firmada -> Solicitud de revisión", round(tiempos_pret.get("promedio_dias_ddjj_a_solicitud_revision", 0), 2)],
        ["Solicitud de revisión -> Aprobado", round(tiempos_pret.get("promedio_dias_revision_a_aprobado", 0), 2)],
    ]
    pdf.add_table(tabla_pret)

    # Ratificación
    tiempos_rat = g(stats, "tiempos.ratificacion", {})
    pdf.section_title("Ratificación")
    tabla_rat = [
        ["Métrica", "Valor"],
        ["Promedio días: 1er mail -> ratificación", round(tiempos_rat.get("promedio_dias_mail_a_ratificacion", 0), 2)],
        ["Ratificaciones pendientes", tiempos_rat.get("ratificaciones_pendientes", 0)],
    ]
    pdf.add_table(tabla_rat)

    # =========================================================
    # Salida
    # =========================================================
    pdf_path = "/tmp/estadisticas_adopciones.pdf"
    pdf.output(pdf_path)
    return FileResponse(pdf_path, filename="estadisticas_adopciones.pdf", media_type="application/pdf")



@estadisticas_router.get("/historial/{login}", response_model=dict, dependencies=[Depends( verify_api_key ), 
                         Depends(require_roles(["administrador", "supervision", "supervisora", "coordinadora"]))])
def get_user_timeline( login: str, db: Session = Depends(get_db) ):
    """
    Devuelve el historial de actividades de un usuario basado en su login (DNI).
    Incluye fechas clave para la línea de tiempo.
    """
    try:

        return {
            "login": "login",
            "fecha_alta": "fecha_alta",
            "fecha_curso_aprobado": "fecha_curso_aprobado",
            "fecha_firma_ddjj": "fecha_firma_ddjj",
            "fecha_aprobacion_doc_personal": "fecha_aprobacion_doc_personal",
            "fecha_presentacion_proyecto": "fecha_presentacion_proyecto",
            "fecha_aprobacion_doc_proyecto": "fecha_aprobacion_doc_proyecto",
            "fecha_asignacion_nro_orden": "fecha_asignacion_nro_orden"
        }

        
        # Buscar usuario
        user = db.query(User).filter(User.login == login).first()
        if not user:
            raise HTTPException(status_code=404, detail="Usuario no encontrado.")

        # Buscar fechas en la tabla User
        fecha_alta = parse_date(user.fecha_alta)
        # fecha_curso_aprobado = parse_date(user.doc_adoptante_curso_aprobado_fecha)
        fecha_firma_ddjj = parse_date(user.doc_adoptante_ddjj_firmada_fecha)
        fecha_aprobacion_doc_personal = parse_date(user.doc_adoptante_estado_fecha)

        # Buscar proyecto asociado (puede tener más de uno, tomamos el más reciente)
        proyecto = (
            db.query(Proyecto)
            .filter((Proyecto.login_1 == login) | (Proyecto.login_2 == login))
            .order_by(Proyecto.fecha_asignacion_nro_orden.desc())  # El más reciente primero
            .first()
        )

        fecha_presentacion_proyecto = parse_date(proyecto.fecha_presentacion) if proyecto else None
        fecha_aprobacion_doc_proyecto = parse_date(proyecto.fecha_aprobacion_doc_proyecto) if proyecto else None
        fecha_asignacion_nro_orden = parse_date(proyecto.fecha_asignacion_nro_orden) if proyecto else None

        # Buscar fechas adicionales en RuaEvento si no están en las otras tablas
        eventos = (
            db.query(RuaEvento.evento_fecha, RuaEvento.evento_detalle)
            .filter(RuaEvento.login == login)
            .all()
        )

        # Mapear eventos específicos a sus fechas
        eventos_dict = {evento.evento_detalle: parse_date(evento.evento_fecha) for evento in eventos}

        # fecha_curso_aprobado = fecha_curso_aprobado or eventos_dict.get("Curso aprobado")
        fecha_firma_ddjj = fecha_firma_ddjj or eventos_dict.get("DDJJ firmada")
        fecha_aprobacion_doc_personal = fecha_aprobacion_doc_personal or eventos_dict.get("Documentación personal aprobada")
        fecha_presentacion_proyecto = fecha_presentacion_proyecto or eventos_dict.get("Proyecto presentado")
        fecha_aprobacion_doc_proyecto = fecha_aprobacion_doc_proyecto or eventos_dict.get("Documentación de proyecto aprobada")
        fecha_asignacion_nro_orden = fecha_asignacion_nro_orden or eventos_dict.get("Número de orden asignado")

        return {
            "login": login,
            "fecha_alta": fecha_alta,
            "fecha_curso_aprobado": fecha_curso_aprobado,
            "fecha_firma_ddjj": fecha_firma_ddjj,
            "fecha_aprobacion_doc_personal": fecha_aprobacion_doc_personal,
            "fecha_presentacion_proyecto": fecha_presentacion_proyecto,
            "fecha_aprobacion_doc_proyecto": fecha_aprobacion_doc_proyecto,
            "fecha_asignacion_nro_orden": fecha_asignacion_nro_orden
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener el historial del usuario: {str(e)}")



# ---------- Utilidades ----------

def _safe_date(d: Optional[date]) -> Optional[str]:
    return d.strftime("%Y-%m-%d") if isinstance(d, (date, datetime)) and d else None


def _autoformat_sheet(ws):
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18

# ---------- Streaming queries (bajo lock de lectura) ----------

def _get_engine():
    # tomamos el engine desde la SessionLocal
    return SessionLocal().bind



PROJECTS_HEADERS = [
    "nro_orden_rua",
    "proyecto_tipo",
    "estado_general",
    "fecha_estado",
    "login_1",
    "login_1_nombre",
    "login_2",
    "login_2_nombre",
    "localidad",
    "Cant. NNA/s",
    "NNA/s",
    "fecha_asignacion_nro_orden",
    "proyecto_id",
]


def _stream_proyectos_rows(conn, estado_filter: Optional[str] = None):
    # Por si el GROUP_CONCAT queda largo y para evitar problemas de collation
    try:
        conn.exec_driver_sql("SET SESSION group_concat_max_len = 8192")
        conn.exec_driver_sql("SET collation_connection = 'utf8mb4_general_ci'")
    except Exception:
        pass

    sql_txt = """
        SELECT
            /* === campos en el orden del header === */
            p.nro_orden_rua,
            p.proyecto_tipo,
            p.estado_general,

            /* fecha del último cambio al estado actual (o fallback) en formato DATE */
            DATE(
              COALESCE(
                (
                  SELECT MAX(he.fecha_hora)
                  FROM proyecto_historial_estado he
                  WHERE he.proyecto_id = p.proyecto_id
                    AND he.estado_nuevo COLLATE utf8mb4_general_ci
                        = p.estado_general COLLATE utf8mb4_general_ci
                ),
                p.ultimo_cambio_de_estado
              )
            ) AS fecha_estado,

            p.login_1,
            TRIM(CONCAT(COALESCE(u1.nombre,''),' ',COALESCE(u1.apellido,''))) AS login_1_nombre,
            p.login_2,
            TRIM(CONCAT(COALESCE(u2.nombre,''),' ',COALESCE(u2.apellido,''))) AS login_2_nombre,

            p.proyecto_localidad AS localidad,

            /* NNA relacionados */
            COALESCE(nn.cant_nnas, 0)    AS cant_nnas,
            COALESCE(nn.nna_nombres, '') AS nna_nombres,

            p.fecha_asignacion_nro_orden,
            p.proyecto_id,

            /* este campo no se exporta, solo por si hiciera falta fallback en Python */
            p.ultimo_cambio_de_estado

        FROM proyecto p
        LEFT JOIN sec_users u1 ON u1.login = p.login_1
        LEFT JOIN sec_users u2 ON u2.login = p.login_2

        /* Subquery que junta los NNA por proyecto */
        LEFT JOIN (
            SELECT
                dp.proyecto_id AS pid,
                COUNT(DISTINCT dn.nna_id) AS cant_nnas,
                GROUP_CONCAT(
                    DISTINCT TRIM(CONCAT(COALESCE(n.nna_nombre,''),' ',COALESCE(n.nna_apellido,'')))
                    ORDER BY n.nna_nombre, n.nna_apellido
                    SEPARATOR ' - '
                ) AS nna_nombres
            FROM detalle_proyectos_en_carpeta dp
            JOIN detalle_nna_en_carpeta dn ON dn.carpeta_id = dp.carpeta_id
            JOIN nna n ON n.nna_id = dn.nna_id
            GROUP BY dp.proyecto_id
        ) nn ON nn.pid = p.proyecto_id
    """

    # Solo proyectos RUA, y opcionalmente por estado
    where_clauses = ["p.ingreso_por COLLATE utf8mb4_general_ci = 'rua'"]
    params = {}
    if estado_filter:
        where_clauses.append("p.estado_general = :estado")
        params["estado"] = estado_filter

    sql_txt += " WHERE " + " AND ".join(where_clauses)

    sql = text(sql_txt).execution_options(stream_results=True)
    cur = conn.execution_options(stream_results=True).execute(sql, params)
    try:
        for row in cur:
            # fecha_estado viene como DATE; formateo seguro a YYYY-MM-DD
            if row.fecha_estado:
                try:
                    fecha_estado_str = row.fecha_estado.strftime("%Y-%m-%d")
                except Exception:
                    fecha_estado_str = str(row.fecha_estado)[:10]
            else:
                # fallback extra a ultimo_cambio_de_estado (no exportado)
                uce = getattr(row, "ultimo_cambio_de_estado", None)
                if uce:
                    try:
                        fecha_estado_str = uce.strftime("%Y-%m-%d")
                    except Exception:
                        try:
                            fecha_estado_str = uce.date().isoformat()
                        except Exception:
                            fecha_estado_str = str(uce)[:10]
                else:
                    fecha_estado_str = None

            yield [
                row.nro_orden_rua,
                row.proyecto_tipo,
                row.estado_general,
                fecha_estado_str,
                row.login_1,
                (row.login_1_nombre or "").strip(),
                row.login_2,
                (row.login_2_nombre or "").strip(),
                row.localidad,
                int(row.cant_nnas or 0),
                row.nna_nombres or "",
                (row.fecha_asignacion_nro_orden.isoformat() if row.fecha_asignacion_nro_orden else None),
                row.proyecto_id,
            ]
    finally:
        cur.close()


NNA_HEADERS = [
    "nombre",
    "apellido",
    "dni",
    "fecha_nacimiento",
    "edad",
    "estado",
    "proyecto",
    "nna_id",
]



# def _stream_nna_rows(conn):
#     sql = text("""
#         SELECT
#             /* === columnas en el orden de NNA_HEADERS === */
#             n.nna_nombre AS nombre,
#             n.nna_apellido AS apellido,
#             n.nna_dni AS dni,
#             n.nna_fecha_nacimiento AS fecha_nacimiento,
#             TIMESTAMPDIFF(YEAR, n.nna_fecha_nacimiento, CURDATE()) AS edad,
#             n.nna_estado AS estado,
#             COALESCE(proj.pretensos, '') AS proyecto,
#             n.nna_id

#         FROM nna n

#         /* Proyecto "más reciente" por NNA (si existe) */
#         LEFT JOIN (
#             SELECT x.nna_id,
#                    x.proyecto_id,
#                    TRIM(
#                      CONCAT(
#                        COALESCE(x.u1_nombre,''),' ',COALESCE(x.u1_apellido,''),
#                        CASE
#                          WHEN x.es_mono = 1 OR x.login_2 IS NULL OR x.login_2 = '' THEN ''
#                          ELSE CONCAT(' - ', COALESCE(x.u2_nombre,''),' ',COALESCE(x.u2_apellido,''))
#                        END
#                      )
#                    ) AS pretensos
#             FROM (
#                 SELECT
#                     dn.nna_id,
#                     p.proyecto_id,
#                     p.login_1,
#                     p.login_2,
#                     CASE WHEN p.proyecto_tipo = 'Monoparental' THEN 1 ELSE 0 END AS es_mono,
#                     u1.nombre AS u1_nombre, u1.apellido AS u1_apellido,
#                     u2.nombre AS u2_nombre, u2.apellido AS u2_apellido,
#                     ROW_NUMBER() OVER (
#                         PARTITION BY dn.nna_id
#                         ORDER BY
#                           COALESCE(p.ultimo_cambio_de_estado, p.fecha_asignacion_nro_orden, '1970-01-01') DESC,
#                           p.proyecto_id DESC
#                     ) AS rn
#                 FROM detalle_nna_en_carpeta dn
#                 JOIN detalle_proyectos_en_carpeta dp ON dp.carpeta_id = dn.carpeta_id
#                 JOIN proyecto p ON p.proyecto_id = dp.proyecto_id
#                 LEFT JOIN sec_users u1 ON u1.login = p.login_1
#                 LEFT JOIN sec_users u2 ON u2.login = p.login_2
#             ) x
#             WHERE x.rn = 1
#         ) proj ON proj.nna_id = n.nna_id
#     """).execution_options(stream_results=True)

#     cur = conn.execution_options(stream_results=True).execute(sql)
#     try:
#         for row in cur:
#             # fecha_nacimiento a ISO si existe
#             if row.fecha_nacimiento:
#                 try:
#                     fecha_nac_str = row.fecha_nacimiento.isoformat()
#                 except Exception:
#                     fecha_nac_str = str(row.fecha_nacimiento)[:10]
#             else:
#                 fecha_nac_str = None

#             # edad: int o None
#             edad_val = int(row.edad) if row.edad is not None else None

#             yield [
#                 row.nombre,
#                 row.apellido,
#                 row.dni,
#                 fecha_nac_str,
#                 edad_val,
#                 row.estado,
#                 row.proyecto or "",
#                 row.nna_id,
#             ]
#     finally:
#         cur.close()


def _stream_nna_rows(conn):
    """
    Genera las filas de la hoja NNA del Excel.
    Mantiene la lógica original para todas las columnas,
    pero usa exactamente la misma lógica de get_nnas para determinar el 'estado'.
    """


    session = SessionLocal()
    try:
        # ---------------------------------------------------------------------
        # 1️⃣ Consulta principal: igual que antes
        # ---------------------------------------------------------------------
        sql = text("""
            SELECT
                n.nna_nombre AS nombre,
                n.nna_apellido AS apellido,
                n.nna_dni AS dni,
                n.nna_fecha_nacimiento AS fecha_nacimiento,
                TIMESTAMPDIFF(YEAR, n.nna_fecha_nacimiento, CURDATE()) AS edad,
                n.nna_estado AS estado_original,
                n.nna_en_convocatoria AS en_conv,
                COALESCE(proj.pretensos, '') AS proyecto,
                n.nna_id
            FROM nna n
            LEFT JOIN (
                SELECT x.nna_id,
                       x.proyecto_id,
                       TRIM(
                         CONCAT(
                           COALESCE(x.u1_nombre,''),' ',COALESCE(x.u1_apellido,''),
                           CASE
                             WHEN x.es_mono = 1 OR x.login_2 IS NULL OR x.login_2 = '' THEN ''
                             ELSE CONCAT(' - ', COALESCE(x.u2_nombre,''),' ',COALESCE(x.u2_apellido,''))
                           END
                         )
                       ) AS pretensos
                FROM (
                    SELECT
                        dn.nna_id,
                        p.proyecto_id,
                        p.login_1,
                        p.login_2,
                        CASE WHEN p.proyecto_tipo = 'Monoparental' THEN 1 ELSE 0 END AS es_mono,
                        u1.nombre AS u1_nombre, u1.apellido AS u1_apellido,
                        u2.nombre AS u2_nombre, u2.apellido AS u2_apellido,
                        ROW_NUMBER() OVER (
                            PARTITION BY dn.nna_id
                            ORDER BY
                              COALESCE(p.ultimo_cambio_de_estado, p.fecha_asignacion_nro_orden, '1970-01-01') DESC,
                              p.proyecto_id DESC
                        ) AS rn
                    FROM detalle_nna_en_carpeta dn
                    JOIN detalle_proyectos_en_carpeta dp ON dp.carpeta_id = dn.carpeta_id
                    JOIN proyecto p ON p.proyecto_id = dp.proyecto_id
                    LEFT JOIN sec_users u1 ON u1.login = p.login_1
                    LEFT JOIN sec_users u2 ON u2.login = p.login_2
                ) x
                WHERE x.rn = 1
            ) proj ON proj.nna_id = n.nna_id
        """).execution_options(stream_results=True)

        cur = conn.execution_options(stream_results=True).execute(sql)

        # ---------------------------------------------------------------------
        # 2️⃣ Pre-cálculos: conjuntos auxiliares (idéntico a get_nnas)
        # ---------------------------------------------------------------------
        # Cargamos todos los IDs que aparecerán en el Excel
        all_nna_ids = [row.nna_id for row in cur]
        cur.close()

        if not all_nna_ids:
            return

        # Reejecutamos para el stream (ya que el cursor fue cerrado)
        cur = conn.execution_options(stream_results=True).execute(sql)

        # Set: NNA en convocatoria
        nna_en_conv_set = set()
        if all_nna_ids:
            nna_en_conv_set = {
                row[0]
                for row in session.query(DetalleNNAEnConvocatoria.nna_id)
                    .filter(DetalleNNAEnConvocatoria.nna_id.in_(all_nna_ids))
                    .distinct()
                    .all()
            }

        # Set: NNA en proyecto
        nna_en_proyecto_set = set()
        if all_nna_ids:
            nna_en_proyecto_set = {
                row[0]
                for row in (
                    session.query(DetalleNNAEnCarpeta.nna_id)
                        .join(
                            DetalleProyectosEnCarpeta,
                            DetalleProyectosEnCarpeta.carpeta_id == DetalleNNAEnCarpeta.carpeta_id
                        )
                        .filter(DetalleNNAEnCarpeta.nna_id.in_(all_nna_ids))
                        .distinct()
                        .all()
                )
            }

        # ---------------------------------------------------------------------
        # 3️⃣ Helper: lógica exacta de get_nnas
        # ---------------------------------------------------------------------
        def traducir_detalle_estado(estado: Optional[str], en_conv: bool) -> str:
            if not estado:
                return ""
            if estado == "disponible":
                return "Esperando flia. en CONV" if en_conv else "Esperando familia"
            mapa = {
                "sin_ficha_sin_sentencia": "Sin ficha ni sentencia",
                "con_ficha_sin_sentencia": "Sólo con ficha",
                "sin_ficha_con_sentencia": "Sólo con sentencia",
                "preparando_carpeta": "Preparando carpeta",
                "enviada_a_juzgado": "Enviado a juzgado",
                "proyecto_seleccionado": "Proyecto seleccionado",
                "vinculacion": "Vinculación",
                "guarda_provisoria": "Guarda provisoria",
                "guarda_confirmada": "Guarda confirmada",
                "adopcion_definitiva": "Adopción definitiva",
                "interrupcion": "Interrupción",
                "mayor_sin_adopcion": "Mayor",
                "no_disponible": "No disponible",
                "en_convocatoria": "Convocatoria",
            }
            return mapa.get(estado, estado)

        # ---------------------------------------------------------------------
        # 4️⃣ Stream final: generar filas con el mismo “detalle_estado”
        # ---------------------------------------------------------------------
        for row in cur:
            # Fecha de nacimiento segura
            if row.fecha_nacimiento:
                try:
                    fecha_nac_str = row.fecha_nacimiento.isoformat()
                except Exception:
                    fecha_nac_str = str(row.fecha_nacimiento)[:10]
            else:
                fecha_nac_str = None

            # Edad
            edad_val = int(row.edad) if row.edad is not None else None

            # Estado (usando sets reales)
            en_conv = (row.nna_id in nna_en_conv_set)
            detalle_estado = traducir_detalle_estado(row.estado_original, en_conv)

            # Resto de columnas igual que antes
            yield [
                row.nombre,
                row.apellido,
                row.dni,
                fecha_nac_str,
                edad_val,
                detalle_estado,   # ← ahora se usa la misma lógica del endpoint
                row.proyecto or "",
                row.nna_id,
            ]

    finally:
        session.close()
        cur.close()




def _build_excel_file(path: str):
    eng = _get_engine()
    with eng.connect() as conn:
        try:
            conn.exec_driver_sql("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED")
            conn.exec_driver_sql("SET SESSION TRANSACTION READ ONLY")
            conn.exec_driver_sql("SET SESSION innodb_lock_wait_timeout=5")
        except Exception:
            pass

        wb = Workbook(write_only=True)

        # --- Hoja Proyectos RUA ---
        ws_p = wb.create_sheet("Proyectos RUA")
        ws_p.append(PROJECTS_HEADERS)
        for row in _stream_proyectos_rows(conn):
            ws_p.append(row)

        # --- Hoja NNA ---
        ws_nna = wb.create_sheet("NNA")
        ws_nna.append(NNA_HEADERS)
        for row in _stream_nna_rows(conn):
            ws_nna.append(row)

        # Guardado en etapa
        root, _ = os.path.splitext(path)
        stage_path = root + ".__stage__.xlsx"
        wb.save(stage_path)

    # Reabrimos para formatear
    wb2 = load_workbook(stage_path)
    for name in ["Proyectos RUA", "NNA"]:
        if name in wb2.sheetnames:
            _autoformat_sheet(wb2[name])
    wb2.save(path)

    try:
        os.remove(stage_path)
    except Exception:
        pass



# ---------- Endpoints Excel ----------
@estadisticas_router.post("/informe_general_excel_job", dependencies=[Depends(verify_api_key),
        Depends(require_roles(["administrador","supervision","supervisora","coordinadora"])) ],)
async def start_informe_general_excel_job(background_tasks: BackgroundTasks):
    # 1) crear job
    job = jobstore_create_job(kind="estadisticas_excel")
    job_id = job["id"]
    out_path = os.path.join(JOBSTORE_EXPORT_DIR, f"estadisticas_{job_id}.xlsx")

    # 2) lanzar tarea en segundo plano
    def _runner():
        try:
            jobstore_update_job(job_id, status="running")
            _build_excel_file(out_path)
            jobstore_update_job(job_id, status="done", file_path=out_path)
        except Exception as e:
            jobstore_update_job(job_id, status="error", error=str(e))

    background_tasks.add_task(_runner)

    # 3) responder al toque
    return {"job_id": job_id, "status": "pending"}


@estadisticas_router.get("/informe_general_excel_job/{job_id}", dependencies=[ Depends(verify_api_key),
        Depends(require_roles(["administrador","supervision","supervisora","coordinadora"]))],)
def get_informe_general_excel_job_status(job_id: str):
    job = jobstore_read_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job no encontrado")
    # No exponemos rutas internas del host si aún no terminó
    return {
        "job_id": job["id"],
        "status": job["status"],
        "error": job.get("error"),
        "file_ready": bool(job.get("file_path") and os.path.exists(job["file_path"])),
    }



@estadisticas_router.get("/informe_general_excel_job/{job_id}/download", dependencies=[ Depends(verify_api_key),
        Depends(require_roles(["administrador","supervision","supervisora","coordinadora"])) ],)
def download_informe_general_excel(job_id: str):
    job = jobstore_read_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job no encontrado")
    if job["status"] != "done" or not job.get("file_path") or not os.path.exists(job["file_path"]):
        raise HTTPException(status_code=409, detail="el archivo todavía no está listo")
    fname = f"estadisticas_adopciones_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return FileResponse(
        job["file_path"],
        filename=fname,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# (Opcional) Endpoint directo "one-shot" (bloquea este worker hasta terminar el Excel)
@estadisticas_router.get("/informe_general_excel", dependencies=[ Depends(verify_api_key),
        Depends(require_roles(["administrador","supervision","supervisora","coordinadora"])) ],)
async def informe_general_excel_directo():
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()
    await run_in_threadpool(_build_excel_file, tmp_path)
    return FileResponse(
        tmp_path,
        filename=f"estadisticas_adopciones_{datetime.now().strftime('%Y%m%d')}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )